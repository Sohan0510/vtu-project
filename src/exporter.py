"""
Excel Exporter
===============
Generates professionally formatted Excel reports from student results.

Sheets:
- "Summary" — One row per student with pass/fail, percentage, GPA
- "Detailed Marks" — One row per subject per student, grouped by semester
- Per-semester sheets — "Sem 1", "Sem 5", etc.
- "Analytics" — Class statistics, toppers, subject-wise pass rates
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from src.config import EXPORTS_DIR


# ── Style Constants ──
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

DATA_FONT = Font(name='Calibri', size=10)
DATA_ALIGNMENT = Alignment(horizontal='center', vertical='center')
LEFT_ALIGNMENT = Alignment(horizontal='left', vertical='center')

PASS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FAIL_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
REVAL_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
TOPPER_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
STATS_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
ALT_ROW_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4')
)

TITLE_FONT = Font(name='Calibri', bold=True, size=16, color='1F3864')
SUBTITLE_FONT = Font(name='Calibri', size=11, color='404040', italic=True)
SECTION_FONT = Font(name='Calibri', bold=True, size=12, color='1F3864')
STATS_FONT = Font(name='Calibri', bold=True, size=10, color='375623')


def _apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _apply_data_style(cell, alignment=None):
    cell.font = DATA_FONT
    cell.alignment = alignment or DATA_ALIGNMENT
    cell.border = THIN_BORDER


def _get_semester(subject_code):
    for ch in subject_code:
        if ch.isdigit():
            return int(ch)
    return 0


def _vtu_grade(total, internals, externals, status):
    """Calculate approximate VTU grade based on percentage."""
    if status == 'F' or status == 'A':
        return status
    if total == 0:
        return '-'
    # Simple grade assignment based on total marks (approximate)
    pct = (total / 100) * 100 if total <= 100 else (total / 150) * 100
    if pct >= 90: return 'O'
    if pct >= 80: return 'A+'
    if pct >= 70: return 'A'
    if pct >= 60: return 'B+'
    if pct >= 50: return 'B'
    if pct >= 40: return 'C'
    return 'P'


def export_to_excel(results, prefix="results"):
    """
    Exports student results to a professionally formatted Excel file.
    
    Returns:
        str: Path to the generated Excel file
    """
    if not results:
        print("[!] No results to export.")
        return None
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_{timestamp}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)
    
    wb = Workbook()
    sorted_results = sorted(results, key=lambda x: x.get("usn", ""))
    
    # ══════════════════════════════════════════
    # SHEET 1: SUMMARY (most useful, first tab)
    # ══════════════════════════════════════════
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_properties.tabColor = "1F3864"
    
    # USN range for title
    usns = [r.get("usn", "") for r in sorted_results if r.get("usn")]
    usn_range = f"{usns[0]} to {usns[-1]}" if usns else "N/A"
    
    ws_summary.merge_cells('A1:K1')
    title_cell = ws_summary['A1']
    title_cell.value = "VTU EXAMINATION RESULTS"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws_summary.merge_cells('A2:K2')
    subtitle_cell = ws_summary['A2']
    subtitle_cell.value = f"USN Range: {usn_range}  |  Students: {len(results)}  |  Generated: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
    subtitle_cell.font = SUBTITLE_FONT
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    summary_headers = [
        "Sl.", "USN", "Student Name", "Semesters",
        "Subjects", "Passed", "Failed",
        "Grand Total", "Percentage", "Overall", "Reval"
    ]
    
    for col, header in enumerate(summary_headers, 1):
        ws_summary.cell(row=4, column=col, value=header)
    _apply_header_style(ws_summary, 4, len(summary_headers))
    
    row_num = 5
    pass_count = 0
    fail_count = 0
    total_marks_all = 0
    max_total = 0
    topper_row = 5
    
    for sl, student in enumerate(sorted_results, 1):
        usn = student.get("usn", "")
        name = student.get("name", "")
        subjects = student.get("subjects", {})
        grand_total = student.get("grand_total", 0)
        
        sems = sorted(set(
            s.get("semester") or _get_semester(code) 
            for code, s in subjects.items()
        ))
        sems_str = ", ".join(str(s) for s in sems)
        
        total_subs = len(subjects)
        passed = sum(1 for s in subjects.values() if s.get("status") == "P")
        failed = sum(1 for s in subjects.values() if s.get("status") in ("F", "A"))
        overall = "PASS" if failed == 0 and total_subs > 0 else "FAIL"
        
        # Calculate percentage (assuming max 100 per subject)
        max_marks = total_subs * 100 if total_subs > 0 else 1
        percentage = round((grand_total / max_marks) * 100, 1) if max_marks > 0 else 0
        
        # Check for reval updates
        has_reval = any(s.get("reval_updated") for s in subjects.values())
        reval_text = "Yes" if has_reval else ""
        
        if overall == "PASS":
            pass_count += 1
        else:
            fail_count += 1
        total_marks_all += grand_total
        
        if grand_total > max_total:
            max_total = grand_total
            topper_row = row_num
        
        row_data = [sl, usn, name, sems_str, total_subs, passed, failed,
                     grand_total, percentage, overall, reval_text]
        
        for col, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_num, column=col, value=value)
            align = LEFT_ALIGNMENT if col == 3 else DATA_ALIGNMENT
            _apply_data_style(cell, align)
            
            if col == 2:  # USN monospace
                cell.font = Font(name='Consolas', size=10, bold=True)
            elif col == 9:  # Percentage
                cell.number_format = '0.0'
            elif col == 10:  # Overall
                cell.fill = PASS_FILL if value == "PASS" else FAIL_FILL
                cell.font = Font(name='Calibri', bold=True, size=10)
            elif col == 11 and value:  # Reval
                cell.fill = REVAL_FILL
            
            # Alternating row colors for readability
            if sl % 2 == 0 and col not in (10, 11):
                cell.fill = ALT_ROW_FILL
        
        row_num += 1
    
    # Highlight topper row
    if max_total > 0:
        for col in range(1, len(summary_headers) + 1):
            cell = ws_summary.cell(row=topper_row, column=col)
            cell.fill = TOPPER_FILL
    
    # ── Statistics row ──
    row_num += 1
    ws_summary.merge_cells(f'A{row_num}:C{row_num}')
    stats_label = ws_summary.cell(row=row_num, column=1, value="CLASS STATISTICS")
    stats_label.font = SECTION_FONT
    
    row_num += 1
    stats_data = [
        ("Total Students", len(sorted_results)),
        ("Passed", pass_count),
        ("Failed", fail_count),
        ("Pass %", f"{round(pass_count / max(len(sorted_results), 1) * 100, 1)}%"),
        ("Class Average", round(total_marks_all / max(len(sorted_results), 1), 1)),
        ("Highest Total", max_total),
    ]
    
    for i, (label, value) in enumerate(stats_data):
        col = i * 2 + 1
        label_cell = ws_summary.cell(row=row_num, column=col, value=label)
        label_cell.font = STATS_FONT
        label_cell.fill = STATS_FILL
        label_cell.border = THIN_BORDER
        value_cell = ws_summary.cell(row=row_num, column=col + 1, value=value)
        value_cell.font = Font(name='Calibri', bold=True, size=10)
        value_cell.fill = STATS_FILL
        value_cell.border = THIN_BORDER
        value_cell.alignment = DATA_ALIGNMENT
    
    summary_widths = [6, 15, 28, 10, 10, 9, 9, 12, 10, 10, 8]
    for i, width in enumerate(summary_widths, 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = width
    
    ws_summary.freeze_panes = 'A5'
    ws_summary.auto_filter.ref = f"A4:K{row_num - 2}"
    
    # ══════════════════════════════════════════
    # SHEET 2: DETAILED MARKS
    # ══════════════════════════════════════════
    ws_detail = wb.create_sheet("Detailed Marks")
    ws_detail.sheet_properties.tabColor = "2E75B6"
    
    ws_detail.merge_cells('A1:J1')
    title_cell = ws_detail['A1']
    title_cell.value = "DETAILED MARKS — ALL SUBJECTS"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    detail_headers = [
        "Sl.", "USN", "Student Name", "Sem", "Subject Code",
        "Subject Name", "Internal", "External", "Total", "Result"
    ]
    
    for col, header in enumerate(detail_headers, 1):
        ws_detail.cell(row=3, column=col, value=header)
    _apply_header_style(ws_detail, 3, len(detail_headers))
    
    row_num = 4
    sl_no = 1
    
    for student in sorted_results:
        usn = student.get("usn", "")
        name = student.get("name", "")
        subjects = student.get("subjects", {})
        
        if not subjects:
            continue
        
        sem_groups = {}
        for sub_code, sub in subjects.items():
            sem = sub.get("semester") or _get_semester(sub_code)
            if sem not in sem_groups:
                sem_groups[sem] = []
            sem_groups[sem].append((sub_code, sub))
        
        for sem in sorted(sem_groups.keys()):
            for sub_code, sub in sorted(sem_groups[sem], key=lambda x: x[0]):
                is_reval = sub.get("reval_updated", False)
                row_data = [
                    sl_no, usn, name, sem, sub_code, sub.get("name", ""),
                    sub.get("internals", 0), sub.get("externals", 0),
                    sub.get("total", 0), sub.get("status", "")
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws_detail.cell(row=row_num, column=col, value=value)
                    align = LEFT_ALIGNMENT if col in (3, 6) else DATA_ALIGNMENT
                    _apply_data_style(cell, align)
                    
                    if col == 2:
                        cell.font = Font(name='Consolas', size=10)
                    elif col == 10:
                        if value == 'P':
                            cell.fill = PASS_FILL
                        elif value in ('F', 'A'):
                            cell.fill = FAIL_FILL
                    
                    if is_reval:
                        cell.fill = REVAL_FILL
                
                sl_no += 1
                row_num += 1
    
    detail_widths = [6, 15, 25, 6, 14, 42, 10, 10, 8, 8]
    for i, width in enumerate(detail_widths, 1):
        ws_detail.column_dimensions[get_column_letter(i)].width = width
    
    ws_detail.freeze_panes = 'A4'
    if row_num > 4:
        ws_detail.auto_filter.ref = f"A3:J{row_num - 1}"
    
    # ══════════════════════════════════════════
    # PER-SEMESTER SHEETS
    # ══════════════════════════════════════════
    all_sems = set()
    for student in sorted_results:
        for code, sub in student.get("subjects", {}).items():
            sem = sub.get("semester") or _get_semester(code)
            all_sems.add(sem)
    
    for sem in sorted(all_sems):
        if sem == 0:
            continue
            
        ws_sem = wb.create_sheet(f"Sem {sem}")
        ws_sem.sheet_properties.tabColor = "70AD47"
        
        ws_sem.merge_cells('A1:I1')
        title_cell = ws_sem['A1']
        title_cell.value = f"SEMESTER {sem} — DETAILED MARKS"
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        sem_headers = [
            "Sl.", "USN", "Student Name", "Subject Code",
            "Subject Name", "Internal", "External", "Total", "Result"
        ]
        for col, header in enumerate(sem_headers, 1):
            ws_sem.cell(row=3, column=col, value=header)
        _apply_header_style(ws_sem, 3, len(sem_headers))
        
        row_num = 4
        sl_no = 1
        
        for student in sorted_results:
            usn = student.get("usn", "")
            name = student.get("name", "")
            subjects = student.get("subjects", {})
            
            sem_subs = {
                code: sub for code, sub in subjects.items()
                if (sub.get("semester") or _get_semester(code)) == sem
            }
            
            if not sem_subs:
                continue
            
            for sub_code in sorted(sem_subs.keys()):
                sub = sem_subs[sub_code]
                is_reval = sub.get("reval_updated", False)
                row_data = [
                    sl_no, usn, name, sub_code, sub.get("name", ""),
                    sub.get("internals", 0), sub.get("externals", 0),
                    sub.get("total", 0), sub.get("status", "")
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws_sem.cell(row=row_num, column=col, value=value)
                    align = LEFT_ALIGNMENT if col in (3, 5) else DATA_ALIGNMENT
                    _apply_data_style(cell, align)
                    
                    if col == 2:
                        cell.font = Font(name='Consolas', size=10)
                    elif col == 9:
                        if value == 'P':
                            cell.fill = PASS_FILL
                        elif value in ('F', 'A'):
                            cell.fill = FAIL_FILL
                    
                    if is_reval:
                        cell.fill = REVAL_FILL
                
                sl_no += 1
                row_num += 1
        
        # Per-semester statistics
        sem_students = set()
        sem_pass = 0
        sem_fail = 0
        for student in sorted_results:
            subjects = student.get("subjects", {})
            sem_subs = {
                code: sub for code, sub in subjects.items()
                if (sub.get("semester") or _get_semester(code)) == sem
            }
            if sem_subs:
                usn = student.get("usn", "")
                sem_students.add(usn)
                has_fail = any(s.get("status") in ("F", "A") for s in sem_subs.values())
                if has_fail:
                    sem_fail += 1
                else:
                    sem_pass += 1
        
        row_num += 1
        ws_sem.merge_cells(f'A{row_num}:C{row_num}')
        stats_cell = ws_sem.cell(row=row_num, column=1, value=f"Sem {sem}: {len(sem_students)} students | {sem_pass} passed | {sem_fail} failed | {round(sem_pass/max(len(sem_students),1)*100,1)}% pass rate")
        stats_cell.font = STATS_FONT
        stats_cell.fill = STATS_FILL
        for col in range(1, 10):
            ws_sem.cell(row=row_num, column=col).border = THIN_BORDER
            ws_sem.cell(row=row_num, column=col).fill = STATS_FILL
        
        sem_widths = [6, 15, 25, 14, 45, 10, 10, 8, 8]
        for i, width in enumerate(sem_widths, 1):
            ws_sem.column_dimensions[get_column_letter(i)].width = width
        
        ws_sem.freeze_panes = 'A4'
        if row_num > 4:
            ws_sem.auto_filter.ref = f"A3:I{row_num - 1}"
    
    # ══════════════════════════════════════════
    # ANALYTICS SHEET
    # ══════════════════════════════════════════
    ws_analytics = wb.create_sheet("Analytics")
    ws_analytics.sheet_properties.tabColor = "FFC000"
    
    ws_analytics.merge_cells('A1:F1')
    title_cell = ws_analytics['A1']
    title_cell.value = "SUBJECT-WISE ANALYSIS"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Collect subject-wise stats
    subject_stats = {}
    for student in sorted_results:
        for code, sub in student.get("subjects", {}).items():
            if code not in subject_stats:
                subject_stats[code] = {
                    "name": sub.get("name", ""),
                    "semester": sub.get("semester", 0),
                    "total_students": 0,
                    "passed": 0,
                    "failed": 0,
                    "total_marks": 0,
                    "max_marks": 0,
                    "min_marks": 999,
                }
            stats = subject_stats[code]
            stats["total_students"] += 1
            total = sub.get("total", 0)
            stats["total_marks"] += total
            if total > stats["max_marks"]:
                stats["max_marks"] = total
            if total < stats["min_marks"]:
                stats["min_marks"] = total
            if sub.get("status") == "P":
                stats["passed"] += 1
            else:
                stats["failed"] += 1
    
    analytics_headers = ["Subject Code", "Subject Name", "Sem", "Students", 
                          "Pass %", "Avg Marks", "Max", "Min"]
    for col, header in enumerate(analytics_headers, 1):
        ws_analytics.cell(row=3, column=col, value=header)
    _apply_header_style(ws_analytics, 3, len(analytics_headers))
    
    row_num = 4
    for code in sorted(subject_stats.keys()):
        stats = subject_stats[code]
        n = stats["total_students"]
        pass_pct = round(stats["passed"] / max(n, 1) * 100, 1)
        avg = round(stats["total_marks"] / max(n, 1), 1)
        
        row_data = [
            code, stats["name"], stats["semester"], n,
            pass_pct, avg, stats["max_marks"],
            stats["min_marks"] if stats["min_marks"] < 999 else 0
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws_analytics.cell(row=row_num, column=col, value=value)
            align = LEFT_ALIGNMENT if col == 2 else DATA_ALIGNMENT
            _apply_data_style(cell, align)
            
            if col == 1:
                cell.font = Font(name='Consolas', size=10, bold=True)
            elif col == 5:  # Pass %
                if value >= 80:
                    cell.fill = PASS_FILL
                elif value < 50:
                    cell.fill = FAIL_FILL
                cell.number_format = '0.0'
        
        row_num += 1
    
    analytics_widths = [14, 45, 6, 10, 10, 10, 8, 8]
    for i, width in enumerate(analytics_widths, 1):
        ws_analytics.column_dimensions[get_column_letter(i)].width = width
    
    ws_analytics.freeze_panes = 'A4'
    if row_num > 4:
        ws_analytics.auto_filter.ref = f"A3:H{row_num - 1}"
    
    # ── SAVE ──
    wb.save(filepath)
    print(f"\n[+] Excel exported: {filepath}")
    return filepath
